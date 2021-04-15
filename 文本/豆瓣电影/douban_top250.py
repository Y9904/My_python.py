import random
import time
import requests
from bs4 import BeautifulSoup
import openpyxl


# 详情页数据
def detaile_movie(href):
    response = requests.get(url=href, headers=headers)
    soup = BeautifulSoup(response.text, "html.parser")

    forms = soup.select("span[property='v:genre']")
    # 合并内容，以  / 分隔
    form = '/'.join(form_data.text for form_data in forms)  # 类型
    # print(f"类型:{form}")

    country_data = soup.select('span.pl')[4]
    # next_sibling  下一个兄弟标签（可以是纯文本）
    country = country_data.next_sibling.strip()  # 制片国家
    # print(f"制片国家:{country}")

    language_data = soup.select('span.pl')[5]
    language = language_data.next_sibling.strip()  # 语言
    # print(f"语言:{language}")

    dates = soup.select("span[property='v:initialReleaseDate']")
    date = '/'.join(date_data.text for date_data in dates)  # 日期
    # print(f"日期:{date}")

    time_movie = soup.select_one("span[property='v:runtime']").attrs["content"]  # 时长
    # print(f"时长:{time_movie}分钟")
    return form, country, language, date, time_movie


def main():
    # 创建xlsx表
    work_book = openpyxl.Workbook()  # 创建工作簿
    sheet = work_book.create_sheet("Top250")  # 创建工作表
    # 写入表头
    col_names = ["排名", "电影名", "链接", "评分", "类型", "国家", "语言", "日期", "时长"]
    for index, name in enumerate(col_names):
        sheet.cell(1, index + 1, name)  # 行  列    内容
    count = 0

    # 翻页
    for page in range(10):
        url = f"https://movie.douban.com/top250?start={page * 25}"
        # 1.获取数据
        response = requests.get(url, headers=headers)
        # 2.解码
        response.encoding = response.apparent_encoding
        # 3.解析数据
        soup = BeautifulSoup(response.text, "html.parser")
        # 4.提取数据
        info_list = soup.select('div.info')
        for info in info_list:
            count += 1
            a_label = info.select_one("div.hd>a")
            href = a_label.attrs["href"]  # 链接
            title = info.select_one("span.title").text  # 电影名
            score = info.select_one("div.bd>div.star>span.rating_num").text  # 评分
            # print(f"电影名:{title}")
            # print(f"链接：{href}")
            # print(f"评分{score}")
            # print("-" * 50)
            movie_list = [count, title, href, score]
            movie_list += detaile_movie(href)  # 列表与元组合并成一个列表
            print(movie_list)
            # 将内容写入 Excel 表中
            for index, item in enumerate(movie_list):
                sheet.cell(count + 1, index + 1, item)
        time.sleep(random.random() * 3 + 5)
    work_book.save("豆瓣电影.xlsx")


if __name__ == '__main__':
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.150 Safari/537.36",
        'Cookie': 'bid=v6UoCJWDsco; douban-fav-remind=1; __yadk_uid=58HhBbi0CfWqGwWSpIcNE97yoRjV6s6O; ll="128464"; _vwo_uuid_v2=D73CD0570C95590A66FA615BA0A2C77CF|bdc8df31aa36f462d599778f6664a485; __utmc=30149280; __utmz=30149280.1613957542.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); _pk_ref.100001.8cb4=%5B%22%22%2C%22%22%2C1613981263%2C%22https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3DT0ClQkD0ocYZbEAJplh8C32tROFMo_NhsDPLRbYEU1jrvHcuJKd4RLTYoFS-phaz%26wd%3D%26eqid%3Dd2114f3f000166cc00000006602ca0f2%22%5D; _pk_ses.100001.8cb4=*; __utma=30149280.146760092.1613957542.1613957542.1613981263.2; __utmt=1; dbcl2="233762439:lkns2EgfjPY"; ck=KEq3; ap_v=0,6.0; __gads=ID=b6ec562f7c3decb9:T=1613537529:S=ALNI_Mb4TY1ve8MyHdKImnrmV1CvnSkEmw; push_noty_num=0; push_doumail_num=0; __utmv=30149280.23376; _pk_id.100001.8cb4=ff0c8d35e160151c.1605798152.8.1613981391.1613957541.; __utmb=30149280.5.10.1613981263'
    }
    main()
