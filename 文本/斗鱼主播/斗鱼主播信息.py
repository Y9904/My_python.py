# 导入模拟HTTP请求的包
import requests
# 导入用于分析数据的包
from bs4 import BeautifulSoup
# 保存到Excel需要的包
import openpyxl


# 获取数据
def get_data(url):
    """对某一个URL进行HTTP请求，并获取返回的文本"""
    # 定义头文件
    head = {
        "user - agent": "Mozilla / 5.0(Windows NT 10.0;WOW64) AppleWebKit / 537.36(KHTML, likeGecko) Chrome / 86.0.4240.111Safari / 537.36"
    }
    response = requests.get(url,headers=head)
    response_text = response.content.decode("utf-8")
    return response_text


# 解析数据
def analyse_data(url):
    """根据文本信息抓取内容"""
    data = get_data(url)
    # 实例化BS4对象
    soup = BeautifulSoup(data, "lxml")
    # print(soup)
    # 第一次筛选
    first_data = soup.find_all("ul", class_="layout-Cover-list")[0]
    # 第二次筛选
    second_data = first_data.find_all("li", class_="layout-Cover-item")
    # 定义一个集合
    data_list = []
    for item in second_data:
        # 定义一个字典
        data_dict = {}
        id = item.find("a").attrs["href"].strip("/")
        data_dict["直播房间号"] = id
        name = item.find("div", class_="DyListCover-userName").text
        data_dict["主播名称"] = name
        follow = item.find("span", class_="DyListCover-hot").text
        data_dict["关注度"] = follow

        data_list.append(data_dict)
    return data_list



"""斗鱼颜值区"""
if __name__ == '__main__':
    url = "https://www.douyu.com/g_yz"
    lis = analyse_data(url)
    #保存
    book = openpyxl.Workbook()
    sheet = book.create_sheet('斗鱼颜值区')


    name_list=['直播房间号','主播名称','关注度']
    for index,data in enumerate(name_list):
        # print(index,data)
        sheet.cell(1, index+1, data)
    for number,dic in enumerate(lis):
        ls=[dic['直播房间号'],dic['主播名称'],dic['关注度']]
        for i,j in enumerate(ls):
            sheet.cell(number+2, i+1, j)

    book.save('斗鱼.xlsx')